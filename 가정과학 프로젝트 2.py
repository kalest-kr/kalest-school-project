import openpyxl
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression

model = LinearRegression()

x_data_list = []
y_data_list = []

exel_file = openpyxl.load_workbook(r"C:\Users\ginok\Downloads\data_set_for_school_porject.xlsx")
selected_sheet = exel_file.active
print(selected_sheet)

n = 11

count = 3

a = 3

def bringing_data(x):
    for count in range(n):
        global a
        for data in x.iter_cols(min_col=a, min_row=3, max_row=3, max_col=a):
            for each in data:
                x_data_list.append(each.value)
        for data in x.iter_cols(min_col=a, min_row=1, max_row=1, max_col=a):
            for each in data:
                y_data_list.append(int(each.value))
        a = a + 4
    print(x_data_list)
    print(y_data_list)
bringing_data(selected_sheet)

plt.bar(y_data_list, x_data_list, width=3)
plt.show()

train_x_data = [[item] for item in x_data_list]
train_y_data = [[item] for item in y_data_list]

print(train_y_data)
print(train_x_data)

model.fit(train_y_data, train_x_data)

plt.scatter(train_y_data, train_x_data)
plt.plot(train_y_data, model.coef_ * train_y_data + model.intercept_)
plt.show()
